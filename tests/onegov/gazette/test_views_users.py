from freezegun import freeze_time
from io import BytesIO
from openpyxl import load_workbook
from tests.onegov.gazette.common import login_admin
from tests.onegov.gazette.common import login_editor_1
from tests.onegov.gazette.common import login_editor_2
from tests.onegov.gazette.common import login_editor_3
from tests.onegov.gazette.common import login_publisher
from webtest import TestApp as Client


def test_view_users(gazette_app):
    client = Client(gazette_app)
    login_admin(client)

    manage = client.get('/users')
    assert 'publisher@example.org' in manage
    assert 'editor1@example.org' in manage
    assert 'editor2@example.org' in manage
    assert 'editor3@example.org' in manage

    # try to add a user with a already taken address
    manage = manage.click("Neu")
    manage.form['role'] = 'editor'
    manage.form['name'] = "New editor"
    manage.form['username'] = "editor1@example.org"
    manage = manage.form.submit()
    assert "Dieser Wert ist bereits vorhanden." in manage

    # add a publisher
    manage = client.get('/users')
    manage = manage.click("Neu")
    manage.form['role'] = 'editor'
    manage.form['name'] = "New user"
    manage.form['username'] = "new_user@example.org"
    manage = manage.form.submit().maybe_follow()
    assert "Benutzer hinzugefügt." in manage
    assert "new_user@example.org" in manage

    assert len(gazette_app.smtp.outbox) == 1
    message = gazette_app.smtp.outbox[0]
    message = message.get_payload(1).get_payload(decode=True)
    message = message.decode('utf-8')
    assert "Benutzerkonto Amtsblattredaktion erstellt" in message

    # make it an editor
    manage = manage.click("Bearbeiten", href="new_user")
    manage.form['role'] = 'member'
    manage = manage.form.submit().maybe_follow()
    assert "Benutzer geändert." in manage

    # try to change the email adress to an already taken one
    manage = manage.click("Bearbeiten", href="new_user")
    manage.form['username'] = 'publisher@example.org'
    manage = manage.form.submit()
    assert "Dieser Wert ist bereits vorhanden." in manage

    # delete user
    manage = client.get('/users').click("Löschen", href="new_user")
    manage = manage.form.submit().maybe_follow()
    assert "Benutzer gelöscht." in manage
    assert "new_user@example.org" not in manage


def test_view_users_permissions(gazette_app):
    client = Client(gazette_app)

    login_admin(client)
    manage = client.get('/users')
    assert "<h3>Redaktoren</h3>" in manage
    assert "<h3>Herausgeber</h3>" in manage
    edit_editor = manage.click("Bearbeiten", href='editor1').request.url
    delete_editor = manage.click("Löschen", href='editor1').request.url
    edit_publisher = manage.click("Bearbeiten", href='publisher').request.url
    delete_publisher = manage.click("Löschen", href='publisher').request.url

    login_publisher(client)
    manage = client.get('/users')
    assert "<h3>Redaktoren</h3>" in manage
    assert "<h3>Herausgeber</h3>" not in manage
    assert manage.click("Bearbeiten", href='editor1').request.url == \
        edit_editor
    assert manage.click("Löschen", href='editor1').request.url == delete_editor
    client.get(edit_publisher, status=403)
    client.get(delete_publisher, status=403)

    login_editor_1(client)
    client.get('/users', status=403)
    client.get(edit_editor, status=403)
    client.get(edit_publisher, status=403)
    client.get(delete_editor, status=403)
    client.get(delete_publisher, status=403)

    login_editor_2(client)
    client.get('/users', status=403)
    client.get(edit_editor, status=403)
    client.get(edit_publisher, status=403)
    client.get(delete_editor, status=403)
    client.get(delete_publisher, status=403)

    login_editor_3(client)
    client.get('/users', status=403)
    client.get(edit_editor, status=403)
    client.get(edit_publisher, status=403)
    client.get(delete_editor, status=403)
    client.get(delete_publisher, status=403)


def test_view_user_sessions(gazette_app):
    admin = Client(gazette_app)
    with freeze_time("2016-06-06 06:06"):
        login_admin(admin)

    client_1 = Client(gazette_app)
    with freeze_time("2017-07-07 07:07"):
        login_editor_1(client_1)

    client_2 = Client(gazette_app)
    with freeze_time("2018-08-08 08:08"):
        login_editor_1(client_2)

    client_1.get('/dashboard')
    client_2.get('/dashboard')

    manage = admin.get('/users/sessions')
    assert '(admin@example.org)' in manage
    assert '(editor1@example.org)' in manage
    assert '2016-06-06T06:06:00' in manage
    assert '2017-07-07T07:07:00' in manage
    assert '2018-08-08T08:08:00' in manage

    manage.click('Sitzungen beenden', href='editor1', index=0).form.submit()

    manage = admin.get('/users/sessions')
    assert '(admin@example.org)' in manage
    assert '(editor1@example.org)' not in manage
    assert '2016-06-06T06:06:00' in manage
    assert '2017-07-07T07:07:00' not in manage
    assert '2018-08-08T08:08:00' not in manage

    client_1.get('/dashboard', status=403)
    client_2.get('/dashboard', status=403)


def test_view_user_delete(gazette_app):
    admin = Client(gazette_app)
    login_admin(admin)

    client_1 = Client(gazette_app)
    login_editor_1(client_1)
    client_1.get('/dashboard')

    client_2 = Client(gazette_app)
    login_editor_1(client_2)
    client_2.get('/dashboard')

    manage = admin.get('/users').click("Löschen", href='editor1')
    manage = manage.form.submit().maybe_follow()
    assert "Benutzer gelöscht." in manage

    client_1.get('/dashboard', status=403)
    client_2.get('/dashboard', status=403)


def test_view_user_modify(gazette_app):
    admin = Client(gazette_app)
    login_admin(admin)

    client_1 = Client(gazette_app)
    login_editor_1(client_1)
    client_1.get('/dashboard')

    client_2 = Client(gazette_app)
    login_editor_1(client_2)
    client_2.get('/dashboard')

    manage = admin.get('/users').click("Bearbeiten", href='editor1')
    manage.form['role'] = 'member'
    manage.form['name'] = "Hans"
    manage = manage.form.submit().maybe_follow()
    assert "Benutzer geändert." in manage

    client_1.get('/dashboard', status=403)
    client_2.get('/dashboard', status=403)


def test_view_users_export(gazette_app):
    admin = Client(gazette_app)
    login_admin(admin)

    result = admin.get('/users').click("Als XLSX herunterladen").form.submit()
    book = load_workbook(BytesIO(result.body))
    assert len(book.worksheets) == 2

    sheet = book['Redaktoren']
    assert sheet.max_column == 3
    assert sheet.max_row == 4

    assert sheet.cell(1, 1).value == 'Gruppe'
    assert sheet.cell(1, 2).value == 'Name'
    assert sheet.cell(1, 3).value == 'E-Mail'

    assert sheet.cell(2, 1).value == 'TestGroup'
    assert sheet.cell(2, 2).value == 'First Editor'
    assert sheet.cell(2, 3).value == 'editor1@example.org'

    assert sheet.cell(3, 1).value == 'TestGroup'
    assert sheet.cell(3, 2).value == 'Second Editor'
    assert sheet.cell(3, 3).value == 'editor2@example.org'

    assert sheet.cell(4, 1).value is None
    assert sheet.cell(4, 2).value == 'Third Editor'
    assert sheet.cell(4, 3).value == 'editor3@example.org'

    sheet = book['Herausgeber']
    assert sheet.max_column == 3
    assert sheet.max_row == 2

    assert sheet.cell(1, 1).value == 'Gruppe'
    assert sheet.cell(1, 2).value == 'Name'
    assert sheet.cell(1, 3).value == 'E-Mail'

    assert sheet.cell(2, 1).value is None
    assert sheet.cell(2, 2).value == 'Publisher'
    assert sheet.cell(2, 3).value == 'publisher@example.org'
